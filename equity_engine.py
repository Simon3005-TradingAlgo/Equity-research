"""
equity_engine.py  ·  Fundamentale Aktienanalyse-Engine (pure Python)
====================================================================
Holt Fundamentaldaten (yfinance), berechnet ALLE Kennzahlen, Bewertung und
ein regelbasiertes Verdikt nativ in Python und exportiert ein formatiertes
Excel. Wird sowohl vom Colab-Notebook als auch von der Streamlit-Web-App
(app.py) genutzt.

    pip install yfinance openpyxl pandas numpy

Kernfunktionen:
    fetch_fundamentals(ticker)          -> dict mit Rohdaten + Meta
    compute(data, **annahmen)           -> dict mit allen Kennzahlen/Bewertung
    analyse(tickers, **annahmen)        -> list[result], inkl. Basket-Peers
    to_excel(results, path)             -> formatiertes Excel
"""
import numpy as np
import pandas as pd

MM = 1_000_000  # Anzeige in Mio.

# ---- yfinance Zeilen-Aliasse (Yahoo aendert Bezeichnungen -> mehrere Versuche) ----
INC = {
    "revenue":    ["Total Revenue", "Operating Revenue"],
    "cogs":       ["Cost Of Revenue", "Reconciled Cost Of Revenue"],
    "ebitda":     ["EBITDA", "Normalized EBITDA"],
    "ebit":       ["EBIT", "Operating Income", "Total Operating Income As Reported"],
    "da":         ["Reconciled Depreciation",
                   "Depreciation And Amortization In Income Statement",
                   "Depreciation Amortization Depletion Income Statement"],
    "interest":   ["Interest Expense", "Interest Expense Non Operating", "Net Interest Income"],
    "taxes":      ["Tax Provision", "Income Tax Expense"],
    "pretax":     ["Pretax Income", "Income Before Tax"],
    "net_income": ["Net Income", "Net Income Common Stockholders",
                   "Net Income From Continuing Operation Net Minority Interest"],
    "shares":     ["Diluted Average Shares", "Basic Average Shares"],
}
BAL = {
    "cash":         ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments"],
    "curr_assets":  ["Current Assets", "Total Current Assets"],
    "total_assets": ["Total Assets"],
    "goodwill":     ["Goodwill And Other Intangible Assets", "Goodwill"],
    "total_debt":   ["Total Debt"],
    "curr_liab":    ["Current Liabilities", "Total Current Liabilities"],
    "equity":       ["Stockholders Equity", "Total Equity Gross Minority Interest", "Common Stock Equity"],
}
CF = {
    "cfo":   ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities"],
    "capex": ["Capital Expenditure", "Capital Expenditures"],
}

# ---------------------------------------------------------------- helpers
def _sdiv(a, b):
    a = np.asarray(a, float); b = np.asarray(b, float)
    return np.divide(a, b, out=np.full(np.broadcast(a, b).shape, np.nan, float),
                     where=(b != 0) & ~np.isnan(b))

def _cagr(arr):
    a = np.asarray(arr, float); idx = np.where(~np.isnan(a))[0]
    if len(idx) < 2: return np.nan
    f, l, n = a[idx[0]], a[idx[-1]], idx[-1] - idx[0]
    if f <= 0 or l <= 0 or n <= 0: return np.nan
    return (l / f) ** (1 / n) - 1

def _last(a):
    a = np.asarray(a, float); v = a[~np.isnan(a)]
    return v[-1] if len(v) else np.nan

def _series(df, names):
    if df is None: return None
    for n in names:
        if n in df.index: return df.loc[n]
    return None

# ---------------------------------------------------------------- fetch
def fetch_fundamentals(ticker, n_years=5):
    import yfinance as yf
    t = yf.Ticker(ticker)
    inc, bal, cf = t.income_stmt, t.balance_sheet, t.cashflow
    if inc is None or inc.empty:
        raise ValueError(f"Keine GUV-Daten fuer '{ticker}'. Ticker/Suffix pruefen oder erneut versuchen.")
    periods = sorted(inc.columns)[-n_years:]          # alt -> neu
    years = [p.year for p in periods]

    def grab(df, table):
        out = {}
        for key, names in table.items():
            s = _series(df, names)
            out[key] = (np.array([float(s.get(p, np.nan)) for p in periods])
                        if s is not None else np.full(len(periods), np.nan))
        return out

    data = {"ticker": ticker, "years": years, "found": {}, "missing": [],
            "period_dates": [pd.Timestamp(p) for p in periods]}
    for df, table, tag in [(inc, INC, "GUV"), (bal, BAL, "Bilanz"), (cf, CF, "Cashflow")]:
        block = grab(df, table)
        for k, names in table.items():
            (data["found"].setdefault(tag, []).append(k)
             if _series(df, names) is not None else data["missing"].append(f"{tag}:{k}"))
        data.update(block)

    name, currency, dps, profile = ticker, "EUR", np.nan, {}

    # robuster Kurs: fast_info -> history -> info
    price = np.nan
    try:
        fi = t.fast_info
        for k in ("last_price", "lastPrice"):
            v = getattr(fi, k, None)
            if v: price = float(v); break
        if (not price or np.isnan(price)):
            try: price = float(fi["lastPrice"])
            except Exception: pass
    except Exception:
        pass
    if not price or np.isnan(price):
        try:
            hist = t.history(period="5d")
            if hist is not None and not hist.empty:
                price = float(hist["Close"].dropna().iloc[-1])
        except Exception:
            pass

    info = {}
    try:
        info = t.info or {}
    except Exception:
        info = {}
    if info:
        name = info.get("longName") or info.get("shortName") or ticker
        currency = info.get("financialCurrency") or info.get("currency") or "EUR"
        if not price or np.isnan(price):
            price = info.get("currentPrice") or info.get("regularMarketPrice") or np.nan
    profile = dict(summary=info.get("longBusinessSummary"), sector=info.get("sector"),
                   industry=info.get("industry"), country=info.get("country"),
                   city=info.get("city"), employees=info.get("fullTimeEmployees"),
                   website=info.get("website"), market_cap=info.get("marketCap"),
                   beta=info.get("beta"), hi52=info.get("fiftyTwoWeekHigh"),
                   lo52=info.get("fiftyTwoWeekLow"),
                   target_mean=info.get("targetMeanPrice"), target_high=info.get("targetHighPrice"),
                   target_low=info.get("targetLowPrice"), target_median=info.get("targetMedianPrice"),
                   n_analysts=info.get("numberOfAnalystOpinions"),
                   rec_key=info.get("recommendationKey"), rec_mean=info.get("recommendationMean"))
    try:
        div = t.dividends
        if div is not None and not div.empty and years:
            dps = float(div[div.index.year == years[-1]].sum())
    except Exception:
        pass
    valid_price = price and not (isinstance(price, float) and np.isnan(price))
    data.update(name=name, currency=currency, price=float(price) if valid_price else np.nan,
                dps=dps, profile=profile)
    return data

# ---------------------------------------------------------------- scoring
def _band(x, cuts, scores, default=3):
    if x is None or (isinstance(x, float) and np.isnan(x)): return default
    for c, s in zip(cuts, scores):
        if x >= c: return s
    return scores[-1]

def _score(roic_wacc, rev_cagr, nd_ebitda, intcov, cashconv, accruals, ebit_m, mos):
    s_prof = _band(roic_wacc, [.08, .05, .02, 0], [5, 4, 3, 2], 1)
    s_grow = _band(rev_cagr,  [.12, .08, .04, 0], [5, 4, 3, 2], 1)
    s_bal = _band(-nd_ebitda if not np.isnan(nd_ebitda) else np.nan,
                  [-1, -2, -3, -4], [5, 4, 3, 2], 1)        # niedriger ND/EBITDA = besser
    if not np.isnan(intcov) and intcov < 4: s_bal = min(s_bal, 2)
    s_cf = _band(cashconv, [1.0, .85, .70, .50], [5, 4, 3, 2], 1)
    if not np.isnan(accruals) and accruals > 0.10: s_cf = min(s_cf, 2)
    s_marg = _band(ebit_m, [.20, .12, .06, 0], [5, 4, 3, 2], 1)
    s_val = _band(mos, [.30, .15, 0, -.10], [5, 4, 3, 2], 1)
    return dict(profitabilitaet=s_prof, wachstum=s_grow, bilanz=s_bal,
                cashflow=s_cf, margen=s_marg, bewertung=s_val)

WEIGHTS = dict(geschaeftsmodell=.12, management=.08, wachstum=.10, profitabilitaet=.20,
               bilanz=.15, cashflow=.15, margen=.05, bewertung=.15)

def _verdict(conv, mos):
    m = 0.0 if (mos is None or np.isnan(mos)) else mos
    if conv >= 3.5 and m >= 0.20: return "KAUFEN"
    if conv >= 3.0 and m >= 0.00: return "HALTEN / BEOBACHTEN"
    if conv < 2.5 or m <= -0.10:  return "MEIDEN / VERKAUFEN"
    return "HALTEN"

# ---------------------------------------------------------------- compute
def compute(data, wacc=0.08, growth=None, terminal=0.025,
            qual_business=3, qual_management=3):
    g = lambda k: np.asarray(data[k], float)
    rev, cogs, da = g("revenue"), g("cogs"), g("da")
    ebit = g("ebit"); ebitda = g("ebitda")
    ebit = np.where(np.isnan(ebit), ebitda - da, ebit)
    ebitda = np.where(np.isnan(ebitda), ebit + da, ebitda)
    interest = np.abs(g("interest"))
    taxes = g("taxes")
    pretax = g("pretax"); pretax = np.where(np.isnan(pretax), ebit - interest, pretax)
    ni = g("net_income"); ni = np.where(np.isnan(ni), pretax - taxes, ni)
    shares = g("shares")
    cash, ca, ta = g("cash"), g("curr_assets"), g("total_assets")
    debt, cl, eq = g("total_debt"), g("curr_liab"), g("equity")
    cfo = g("cfo"); capex = np.abs(g("capex")); fcf = cfo - capex

    gross = rev - cogs
    net_debt = debt - cash
    inv_cap = debt + eq - cash
    eff_tax = np.clip(_sdiv(taxes, pretax), 0, 0.5)
    nopat = ebit * (1 - np.where(np.isnan(eff_tax), 0.25, eff_tax))
    eps = _sdiv(ni, shares); fcf_ps = _sdiv(fcf, shares)

    R = {
        "Bruttomarge": _sdiv(gross, rev), "EBITDA-Marge": _sdiv(ebitda, rev),
        "EBIT-Marge": _sdiv(ebit, rev), "Nettomarge": _sdiv(ni, rev),
        "FCF-Marge": _sdiv(fcf, rev),
        "ROE": _sdiv(ni, eq), "ROA": _sdiv(ni, ta), "ROIC": _sdiv(nopat, inv_cap),
        "DuPont Nettomarge": _sdiv(ni, rev), "Kapitalumschlag": _sdiv(rev, ta),
        "EK-Multiplikator": _sdiv(ta, eq),
        "Cash Conversion (CFO/NI)": _sdiv(cfo, ni), "FCF/NI": _sdiv(fcf, ni),
        "Accruals-Ratio": _sdiv(ni - cfo, ta),
        "Nettoverschuldung/EBITDA": _sdiv(net_debt, ebitda),
        "Zinsdeckung": _sdiv(ebit, interest), "Verschuldungsgrad": _sdiv(debt, eq),
        "Current Ratio": _sdiv(ca, cl),
    }
    roic = R["ROIC"]; R["ROIC - WACC"] = roic - wacc

    rev_cagr, eps_cagr, fcf_cagr = _cagr(rev), _cagr(eps), _cagr(fcf)
    if growth is None:
        base = fcf_cagr if not np.isnan(fcf_cagr) else rev_cagr
        growth = float(np.clip(base, 0.0, 0.10)) if not np.isnan(base) else 0.04

    price = data.get("price", np.nan); sh = _last(shares)
    nd = _last(net_debt); mcap = price * sh; ev = mcap + nd
    L = dict(price=price, shares=sh, mktcap=mcap, ev=ev,
             eps=_last(eps), ebit=_last(ebit), ebitda=_last(ebitda),
             fcf=_last(fcf), ni=_last(ni), equity=_last(eq),
             bvps=_sdiv(_last(eq), sh), dps=data.get("dps", np.nan))
    mult = dict(
        pe=_sdiv(price, L["eps"]), ev_ebit=_sdiv(ev, L["ebit"]),
        ev_ebitda=_sdiv(ev, L["ebitda"]), ev_fcf=_sdiv(ev, L["fcf"]),
        pb=_sdiv(price, L["bvps"]), fcf_yield=_sdiv(L["fcf"], mcap),
        div_yield=_sdiv(L["dps"], price),
    )

    # DCF (FCFF, 5J explizit + Gordon-Terminal)
    fair = mos = np.nan; dcf = {}
    if not np.isnan(L["fcf"]) and wacc > terminal:
        proj = [L["fcf"] * (1 + growth) ** k for k in range(1, 6)]
        dfac = [(1 + wacc) ** -k for k in range(1, 6)]
        pv = [f * d for f, d in zip(proj, dfac)]
        tv = proj[-1] * (1 + terminal) / (wacc - terminal)
        pv_tv = tv * dfac[-1]
        ev_dcf = sum(pv) + pv_tv
        eqv = ev_dcf - nd
        fair = eqv / sh if sh else np.nan
        mos = (fair / price - 1) if (price and not np.isnan(price)) else np.nan
        dcf = dict(growth=growth, terminal=terminal, wacc=wacc, base_fcf=L["fcf"],
                   ev_dcf=ev_dcf, equity_value=eqv, fair_value=fair, mos=mos)
    rev_g = ((ev * wacc - L["fcf"]) / (ev + L["fcf"])
             if (not np.isnan(ev) and not np.isnan(L["fcf"]) and (ev + L["fcf"]) != 0) else np.nan)

    sc = _score(_last(R["ROIC - WACC"]), rev_cagr, _last(R["Nettoverschuldung/EBITDA"]),
                _last(R["Zinsdeckung"]), _last(R["Cash Conversion (CFO/NI)"]),
                _last(R["Accruals-Ratio"]), _last(R["EBIT-Marge"]), mos)
    sc["geschaeftsmodell"] = qual_business; sc["management"] = qual_management
    conv = sum(WEIGHTS[k] * sc[k] for k in WEIGHTS)
    verdict = _verdict(conv, mos)

    fin = pd.DataFrame({
        "Umsatz": rev / MM, "Bruttogewinn": gross / MM, "EBITDA": ebitda / MM,
        "EBIT": ebit / MM, "Nettoergebnis": ni / MM, "EPS": eps,
        "Free Cashflow": fcf / MM, "Nettoverschuldung": net_debt / MM,
        "Eigenkapital": eq / MM, "Inv. Kapital": inv_cap / MM,
    }, index=data["years"]).T
    rat = pd.DataFrame(R, index=data["years"]).T

    series = dict(dates=data.get("period_dates"), eps=eps, ebit=ebit, ebitda=ebitda,
                  net_debt=net_debt, shares=shares)

    return dict(ticker=data["ticker"], name=data.get("name"), currency=data.get("currency"),
                years=data["years"], financials=fin, ratios=rat, latest=L, multiples=mult,
                dcf=dcf, reverse_growth=rev_g, scores=sc, weights=WEIGHTS, series=series,
                conviction=conv, verdict=verdict,
                peer=dict(med_ev_ebit=np.nan, med_pe=np.nan, impl_ev_ebit=np.nan, impl_pe=np.nan),
                headline=dict(
                    rev_cagr=rev_cagr, eps_cagr=eps_cagr, fcf_cagr=fcf_cagr,
                    roic=_last(roic), roic_wacc=_last(R["ROIC - WACC"]),
                    ebit_margin=_last(R["EBIT-Marge"]), fcf_margin=_last(R["FCF-Marge"]),
                    nd_ebitda=_last(R["Nettoverschuldung/EBITDA"]),
                    cash_conv=_last(R["Cash Conversion (CFO/NI)"]),
                    fair_value=fair, mos=mos),
                profile=data.get("profile", {}), found=data.get("found"), missing=data.get("missing"))

# ---------------------------------------------------------------- analyse (multi)
def analyse(tickers, wacc=0.08, growth=None, terminal=0.025,
            qual_business=3, qual_management=3, verbose=True):
    if isinstance(tickers, str):
        tickers = [x.strip() for x in tickers.replace(";", ",").split(",") if x.strip()]
    results, errors = [], []
    for tk in tickers:
        try:
            d = fetch_fundamentals(tk)
            r = compute(d, wacc=wacc, growth=growth, terminal=terminal,
                        qual_business=qual_business, qual_management=qual_management)
            results.append(r)
            if verbose:
                print(f"  OK  {tk:<10} {r['name'][:34]:<34} Verdikt: {r['verdict']}")
                if r["missing"]:
                    print(f"      ! fehlend: {', '.join(r['missing'])}")
        except Exception as e:
            errors.append((tk, str(e)))
            if verbose: print(f"  --  {tk:<10} FEHLER: {e}")
    # Basket-Peers: relative Bewertung gegen Median der eingegebenen Titel
    ev_ebits = [r["multiples"]["ev_ebit"] for r in results if not np.isnan(r["multiples"]["ev_ebit"])]
    pes = [r["multiples"]["pe"] for r in results if not np.isnan(r["multiples"]["pe"])]
    med_ev_ebit = float(np.median(ev_ebits)) if ev_ebits else np.nan
    med_pe = float(np.median(pes)) if pes else np.nan
    for r in results:
        nd = r["financials"].loc["Nettoverschuldung"].iloc[-1] * MM
        ebit = r["latest"]["ebit"]; sh = r["latest"]["shares"]; eps = r["latest"]["eps"]
        r["peer"] = dict(med_ev_ebit=med_ev_ebit, med_pe=med_pe,
                         impl_ev_ebit=((med_ev_ebit * ebit - nd) / sh
                                       if sh and not np.isnan(med_ev_ebit) else np.nan),
                         impl_pe=med_pe * eps if not np.isnan(med_pe) else np.nan)
    return results, errors

# ---------------------------------------------------------------- Excel export
def to_excel(results, path="Equity_Analyse.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    F = "Arial"; HDR = PatternFill("solid", fgColor="1F3864"); SEC = PatternFill("solid", fgColor="D9E1F2")
    GRN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
    YEL = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BFBFBF"); BORD = Border(thin, thin, thin, thin)
    PCT = "0.0%"; MUL = '0.0"x"'; CUR = '#,##0;(#,##0);"-"'; PS = '#,##0.00'
    def st(cell, v=None, *, b=False, col="000000", fill=None, fmt=None, al=None, sz=10):
        if v is not None: cell.value = v
        cell.font = Font(name=F, bold=b, color=col, size=sz)
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        if al: cell.alignment = Alignment(horizontal=al, vertical="center")

    def num(x): return None if (x is None or (isinstance(x, float) and np.isnan(x))) else float(x)

    wb = Workbook(); ws = wb.active; ws.title = "Vergleich"; ws.sheet_view.showGridLines = False
    cols = [("Ticker", 12, None), ("Name", 30, None), ("Verdikt", 20, None),
            ("Conviction", 11, "0.0"), ("Kurs", 10, PS), ("Fair Value", 11, PS),
            ("Sicherheitsmarge", 14, PCT), ("ROIC", 9, PCT), ("ROIC-WACC", 11, PCT),
            ("ND/EBITDA", 11, MUL), ("EBIT-Marge", 11, PCT), ("FCF-Marge", 11, PCT),
            ("Cash Conv", 11, PCT), ("KGV", 9, MUL), ("EV/EBIT", 10, MUL),
            ("FCF-Rendite", 11, PCT), ("Rev-CAGR", 10, PCT)]
    st(ws.cell(1, 1), "EQUITY-ANALYSE · VERGLEICH", b=True, col="FFFFFF", fill=HDR, sz=13)
    for i in range(2, len(cols) + 1): ws.cell(1, i).fill = HDR
    for j, (h, w, _) in enumerate(cols, 1):
        st(ws.cell(2, j), h, b=True, fill=SEC, al="center"); ws.column_dimensions[ws.cell(2, j).column_letter].width = w
    for i, r in enumerate(results, 3):
        h = r["headline"]
        row = [r["ticker"], (r["name"] or "")[:30], r["verdict"], num(r["conviction"]),
               num(r["latest"]["price"]), num(h["fair_value"]), num(h["mos"]), num(h["roic"]),
               num(h["roic_wacc"]), num(h["nd_ebitda"]), num(h["ebit_margin"]), num(h["fcf_margin"]),
               num(h["cash_conv"]), num(r["multiples"]["pe"]), num(r["multiples"]["ev_ebit"]),
               num(r["multiples"]["fcf_yield"]), num(h["rev_cagr"])]
        for j, (v, (_, _, fmt)) in enumerate(zip(row, cols), 1):
            st(ws.cell(i, j), v, fmt=fmt, al=("left" if j <= 3 else "right"))
        vc = ws.cell(i, 3)
        vc.fill = GRN if "KAUFEN" in r["verdict"] else (RED if "MEIDEN" in r["verdict"] else YEL)
        vc.font = Font(name=F, bold=True, size=10)
    # Chart: ROIC-WACC Spread
    n = len(results)
    if n:
        chart = BarChart(); chart.title = "ROIC - WACC Spread"; chart.height = 7; chart.width = 15
        chart.add_data(Reference(ws, min_col=9, min_row=2, max_row=2 + n), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=2 + n))
        ws.add_chart(chart, f"A{5 + n}")

    # Detail je Ticker
    for r in results:
        s = wb.create_sheet(r["ticker"][:28]); s.sheet_view.showGridLines = False
        s.column_dimensions["A"].width = 30
        yrs = r["years"]
        for k, _ in enumerate(yrs): s.column_dimensions[chr(66 + k)].width = 12
        st(s.cell(1, 1), f"{r['name']}  ({r['ticker']})", b=True, col="FFFFFF", fill=HDR, sz=12)
        for c in range(2, len(yrs) + 2): s.cell(1, c).fill = HDR
        st(s.cell(2, 1), f"Verdikt: {r['verdict']}   ·   Conviction {r['conviction']:.1f}/5   ·   "
                         f"Fair Value {r['headline']['fair_value']:.2f}   ·   "
                         f"Marge {0 if np.isnan(r['headline']['mos']) else r['headline']['mos']*100:.0f}%",
           b=True, fill=YEL)
        row = 4
        def block(title, df, fmt):
            nonlocal row
            st(s.cell(row, 1), title, b=True, fill=SEC)
            for c, y in enumerate(yrs, 2): st(s.cell(row, c), str(y), b=True, fill=SEC, al="right")
            row += 1
            for idx, vals in df.iterrows():
                st(s.cell(row, 1), idx)
                for c, v in enumerate(vals.values, 2): st(s.cell(row, c), num(v), fmt=fmt, al="right")
                row += 1
            row += 1
        block("Finanzdaten (Mio.)", r["financials"], CUR)
        block("Kennzahlen", r["ratios"], PCT)
        # Bewertung
        st(s.cell(row, 1), "Bewertung", b=True, fill=SEC); row += 1
        m = r["multiples"]; d = r["dcf"]; pr = r["peer"]
        valrows = [("KGV", m["pe"], MUL), ("EV/EBIT", m["ev_ebit"], MUL), ("EV/EBITDA", m["ev_ebitda"], MUL),
                   ("EV/FCF", m["ev_fcf"], MUL), ("FCF-Rendite", m["fcf_yield"], PCT),
                   ("Dividendenrendite", m["div_yield"], PCT),
                   ("DCF Fair Value", d.get("fair_value", np.nan), PS),
                   ("Sicherheitsmarge", d.get("mos", np.nan), PCT),
                   ("DCF Wachstum (Annahme)", d.get("growth", np.nan), PCT),
                   ("WACC (Annahme)", d.get("wacc", np.nan), PCT),
                   ("Reverse-DCF impl. Wachstum", r["reverse_growth"], PCT),
                   ("Peer impl. Wert (EV/EBIT)", pr["impl_ev_ebit"], PS),
                   ("Peer impl. Wert (KGV)", pr["impl_pe"], PS)]
        for lab, v, fmt in valrows:
            st(s.cell(row, 1), lab); st(s.cell(row, 2), num(v), fmt=fmt, al="right"); row += 1
        row += 1
        # Scorecard
        st(s.cell(row, 1), "Scorecard", b=True, fill=SEC)
        st(s.cell(row, 2), "Gewicht", b=True, fill=SEC, al="right")
        st(s.cell(row, 3), "Score", b=True, fill=SEC, al="right"); row += 1
        labels = dict(geschaeftsmodell="Geschaeftsmodell*", management="Management*",
                      wachstum="Wachstum", profitabilitaet="Profitabilitaet (ROIC>WACC)",
                      bilanz="Bilanz & Verschuldung", cashflow="Cashflow-Qualitaet",
                      margen="Margen", bewertung="Bewertung & Marge")
        for k in r["weights"]:
            st(s.cell(row, 1), labels[k]); st(s.cell(row, 2), r["weights"][k], fmt=PCT, al="right")
            st(s.cell(row, 3), r["scores"][k], al="right"); row += 1
        st(s.cell(row, 1), "Gesamt (Conviction)", b=True)
        st(s.cell(row, 3), round(r["conviction"], 1), b=True, al="right"); row += 2
        st(s.cell(row, 1), "* qualitativ, Default 3 - manuell anpassen (qual_business/qual_management).",
           col="808080", sz=8)
    wb.save(path)
    return path

# ---------------------------------------------------------------- self-test
def _synthetic():
    y = lambda v: np.array(v, float) * MM
    d = dict(ticker="MUST", name="Muster AG", currency="EUR", years=[2020,2021,2022,2023,2024,2025],
             revenue=y([4000,4350,4750,5150,5600,6100]), cogs=y([2360,2545,2755,2960,3190,3450]),
             ebitda=y([960,1065,1190,1300,1430,1570]), ebit=np.full(6,np.nan),
             da=y([240,255,275,295,315,340]), interest=y([85,80,75,70,66,60]),
             taxes=y([159,183,210,234,262,293]), pretax=np.full(6,np.nan), net_income=np.full(6,np.nan),
             shares=np.array([510,505,500,495,490,485])*MM,
             cash=y([600,640,690,720,800,900]), curr_assets=y([1800,1900,2050,2200,2380,2560]),
             total_assets=y([7800,8050,8350,8650,9000,9400]), goodwill=y([2200,2180,2160,2300,2280,2260]),
             total_debt=y([1900,1820,1760,1700,1640,1560]), curr_liab=y([1500,1560,1640,1720,1820,1920]),
             equity=y([3300,3560,3850,4180,4560,4990]), cfo=y([690,760,860,940,1040,1150]),
             capex=y([-280,-300,-330,-360,-390,-420]), price=28.0, dps=0.88, found={}, missing=[],
             period_dates=[pd.Timestamp(f"{yr}-12-31") for yr in [2020,2021,2022,2023,2024,2025]])
    return d

if __name__ == "__main__":
    r = compute(_synthetic())
    print("Verdikt:", r["verdict"], "| Conviction:", round(r["conviction"],2),
          "| ROIC:", round(r["headline"]["roic"],3), "| MoS:", round(r["headline"]["mos"],3))
    to_excel([r], "selftest.xlsx"); print("Excel ok")
