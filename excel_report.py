"""
excel_report.py - professioneller Excel-Bericht (Einzeltitel, ohne Anlageempfehlung)
====================================================================================
build_excel(main, peer_rs, models, blended, consensus, wacc) -> bytes (XLSX)

Blaetter: Uebersicht | Abschluesse | Kennzahlen | Bewertung | Dividende |
          Qualitaet | (Peers). Spiegelt den vollen Dashboard-Inhalt mit
          Tabellen, Charts und DCF-Sensitivitaet. Keine Anlageempfehlung.
"""
from io import BytesIO
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule

MM = 1_000_000; F = "Arial"
HDR = PatternFill("solid", fgColor="1F3864"); SEC = PatternFill("solid", fgColor="D9E1F2")
GRN = PatternFill("solid", fgColor="C6EFCE"); REDF = PatternFill("solid", fgColor="FFC7CE")
YEL = PatternFill("solid", fgColor="FFF2CC"); LITE = PatternFill("solid", fgColor="F2F5FA")
thin = Side(style="thin", color="C8D0DC"); BORD = Border(thin, thin, thin, thin)
PCT = "0.0%"; MUL = '0.0"x"'; CUR = '#,##0;(#,##0);"-"'; PS = "#,##0.00"; UP = "+0.0%;-0.0%;0.0%"

def _n(x): return None if x is None or (isinstance(x, float) and np.isnan(x)) else float(x)
def _arr(df, name): return df.loc[name].values.astype(float) if name in df.index else None

def C(ws, coord, v=None, *, b=False, col="000000", fill=None, fmt=None, al=None, sz=10, wrap=False, bd=False):
    c = ws[coord]
    if v is not None: c.value = v
    c.font = Font(name=F, bold=b, color=col, size=sz)
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if al or wrap: c.alignment = Alignment(horizontal=al, vertical="center", wrap_text=wrap)
    if bd: c.border = BORD
    return c

def title(ws, coord, span, text):
    C(ws, coord, text, b=True, col="FFFFFF", fill=HDR, sz=13)
    for col in range(ws[coord].column, ws[span].column + 1): ws.cell(ws[coord].row, col).fill = HDR

def section(ws, row, last, text):
    C(ws, f"A{row}", text, b=True, sz=11, fill=SEC)
    for col in range(1, ws[f"{last}{row}"].column + 1): ws.cell(row, col).fill = SEC

def _statement_block(ws, row, dfrow, header_years, ccy, twocol=("EPS", "FCF je", "Aktien")):
    """Schreibt einen Abschluss-Block ab Zeile row; gibt naechste freie Zeile."""
    C(ws, f"A{row}", header_years[0], b=True, fill=SEC, bd=True)
    for c, y in enumerate(header_years[1], 2): C(ws, f"{chr(64+c)}{row}", str(y), b=True, fill=SEC, al="right", bd=True)
    row += 1
    for name in dfrow.index:
        C(ws, f"A{row}", name, bd=True)
        fmt = PS if any(str(name).startswith(t) for t in twocol) else CUR
        for c, v in enumerate(dfrow.loc[name].values, 2):
            C(ws, f"{chr(64+c)}{row}", _n(v), fmt=fmt, al="right", bd=True)
        row += 1
    return row + 1

# --------------------------------------------------------------- Uebersicht
def _overview(ws, main, models, blended, consensus, ccy):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 30
    for col in "BCDEF": ws.column_dimensions[col].width = 15
    title(ws, "A1", "F1", "EQUITY RESEARCH · UEBERSICHT")
    C(ws, "A2", main["name"], b=True, sz=14); ws.merge_cells("A2:D2")
    C(ws, "E2", f"{main['ticker']} · {ccy}", al="right", col="5A5A5A")

    price = main["latest"]["price"]; h, m = main["headline"], main["multiples"]; p = main.get("profile") or {}
    vals = [v for v in models.values() if v is not None and not np.isnan(v)]
    lo, hi = (min(vals), max(vals)) if vals else (np.nan, np.nan)
    mos = (blended / price - 1) if price and not np.isnan(price) and not np.isnan(blended) else np.nan
    C(ws, "A4", "Aktueller Kurs", b=True, fill=SEC, bd=True); C(ws, "B4", _n(price), b=True, fmt=PS, al="right", bd=True)
    C(ws, "C4", "Qualitaets-Score", b=True, fill=SEC, al="right", bd=True); C(ws, "D4", round(main["conviction"], 1), b=True, fmt='0.0" / 5,0"', al="right", bd=True)
    C(ws, "A5", "Bewertungsspanne (Methoden)", b=True, fill=SEC, bd=True)
    C(ws, "B5", f"{_n(lo):,.2f}  bis  {_n(hi):,.2f}" if vals else "n/v", b=True, al="right", bd=True); ws.merge_cells("B5:C5")
    C(ws, "D5", "Median / Auf-Abschlag", b=True, fill=SEC, al="right", bd=True)
    C(ws, "E5", _n(blended), b=True, fmt=PS, al="right", bd=True); C(ws, "F5", _n(mos), b=True, fmt=UP, al="right", bd=True)

    section(ws, 7, "F", "Unternehmensprofil")
    prof = [("Sektor", p.get("sector")), ("Industrie", p.get("industry")),
            ("Sitz", f"{(p.get('city')+', ') if p.get('city') else ''}{p.get('country') or ''}".strip(", ") or None),
            ("Mitarbeiter", p.get("employees")), ("Marktkap. (Mio.)", (p.get("market_cap") or 0) / MM if p.get("market_cap") else None),
            ("Beta", p.get("beta")), ("Website", p.get("website"))]
    r = 8
    for i in range(0, len(prof), 2):
        for j in range(2):
            if i + j < len(prof):
                lab, val = prof[i + j]; cc = "A" if j == 0 else "D"; vc = "B" if j == 0 else "E"
                C(ws, f"{cc}{r}", lab, bd=True)
                if isinstance(val, (int, float)) and val is not None:
                    C(ws, f"{vc}{r}", float(val), fmt=("#,##0" if lab.startswith("Markt") or lab == "Mitarbeiter" else "0.00"), al="right", bd=True)
                else:
                    C(ws, f"{vc}{r}", val or "n/v", al="right", bd=True)
        r += 1
    if p.get("summary"):
        section(ws, r + 1, "F", "Geschaeftsbeschreibung")
        C(ws, f"A{r+2}", p["summary"][:1800], wrap=True, sz=9); 
        ws.merge_cells(f"A{r+2}:F{r+7}")
        for rr in range(r + 2, r + 8): ws.row_dimensions[rr].height = 26
        r += 8

    section(ws, r + 1, "F", "Kennzahlen-Snapshot (letztes GJ)")
    kpis = [("ROIC", _n(h["roic"]), PCT), ("ROIC - WACC", _n(h["roic_wacc"]), PCT), ("EBIT-Marge", _n(h["ebit_margin"]), PCT),
            ("FCF-Marge", _n(h["fcf_margin"]), PCT), ("Umsatz-CAGR", _n(h["rev_cagr"]), PCT), ("Net Debt/EBITDA", _n(h["nd_ebitda"]), MUL),
            ("Cash Conversion", _n(h["cash_conv"]), PCT), ("KGV", _n(m["pe"]), MUL), ("EV/EBIT", _n(m["ev_ebit"]), MUL),
            ("EV/EBITDA", _n(m["ev_ebitda"]), MUL), ("FCF-Rendite", _n(m["fcf_yield"]), PCT), ("Div.-Rendite", _n(m["div_yield"]), PCT)]
    r += 2
    for i in range(0, len(kpis), 2):
        for j in range(2):
            if i + j < len(kpis):
                lab, val, fmt = kpis[i + j]
                C(ws, f"{'A' if j==0 else 'D'}{r}", lab, bd=True)
                C(ws, f"{'B' if j==0 else 'E'}{r}", val, fmt=fmt, al="right", bd=True)
        r += 1

    section(ws, r + 1, "F", "Bewertungsmethoden")
    hr = r + 2
    C(ws, f"A{hr}", "Methode", b=True, fill=LITE, bd=True); C(ws, f"B{hr}", "Fairer Wert", b=True, fill=LITE, al="right", bd=True)
    C(ws, f"C{hr}", "Auf-/Abschlag", b=True, fill=LITE, al="right", bd=True)
    rr = hr + 1
    for lab, val in models.items():
        if val is None or np.isnan(val): continue
        up = (val / price - 1) if price and not np.isnan(price) else np.nan
        C(ws, f"A{rr}", lab, bd=True); C(ws, f"B{rr}", _n(val), fmt=PS, al="right", bd=True)
        c = C(ws, f"C{rr}", _n(up), fmt=UP, al="right", bd=True)
        if up is not None and not np.isnan(up): c.fill = GRN if up >= 0 else REDF
        rr += 1
    C(ws, f"A{rr}", "Median", b=True, bd=True); C(ws, f"B{rr}", _n(blended), b=True, fmt=PS, al="right", bd=True)
    C(ws, f"C{rr}", _n(mos), b=True, fmt=UP, al="right", bd=True)
    C(ws, f"A{rr+1}", "Aktueller Kurs", b=True, fill=YEL, bd=True); C(ws, f"B{rr+1}", _n(price), b=True, fill=YEL, fmt=PS, al="right", bd=True)
    ch = BarChart(); ch.type = "bar"; ch.title = f"Fairer Wert je Methode ({ccy})"; ch.height = 7; ch.width = 12; ch.legend = None
    ch.add_data(Reference(ws, min_col=2, min_row=hr + 1, max_row=rr + 1)); ch.set_categories(Reference(ws, min_col=1, min_row=hr + 1, max_row=rr + 1))
    ws.add_chart(ch, f"D{r+2}")
    C(ws, f"A{rr+3}", "Nur zu Informationszwecken - keine Anlageempfehlung. Quelle: Yahoo Finance / FMP (zu pruefen).", col="5A5A5A", sz=9)
    ws.merge_cells(f"A{rr+3}:F{rr+3}")

# --------------------------------------------------------------- Abschluesse
def _statements(ws, main, ccy):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 30
    yy = main["years"]
    for k in range(len(yy)): ws.column_dimensions[chr(66 + k)].width = 14
    title(ws, "A1", f"{chr(65+len(yy))}1", f"ABSCHLUESSE (Mio. {ccy})")
    stm = main["statements"]
    row = 3
    section(ws, row, chr(65 + len(yy)), "Gewinn- und Verlustrechnung"); row += 1
    row = _statement_block(ws, row, stm["income"], ("Position", yy), ccy)
    section(ws, row, chr(65 + len(yy)), "Bilanz"); row += 1
    row = _statement_block(ws, row, stm["balance"], ("Position", yy), ccy)
    section(ws, row, chr(65 + len(yy)), "Kapitalflussrechnung"); row += 1
    row = _statement_block(ws, row, stm["cashflow"], ("Position", yy), ccy)
    C(ws, f"A{row}", "Negative Werte (in Klammern) = Aufwendungen/Investitionen. Free Cashflow = CFO - CapEx.", col="5A5A5A", sz=9)

# --------------------------------------------------------------- Kennzahlen
def _ratios(ws, main):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 30
    yrs = main["years"]
    for k in range(len(yrs)): ws.column_dimensions[chr(66 + k)].width = 13
    title(ws, "A1", f"{chr(65+len(yrs))}1", "KENNZAHLEN")
    groups = {"Margen": ["Bruttomarge", "EBITDA-Marge", "EBIT-Marge", "Nettomarge", "FCF-Marge"],
              "Kapitalrendite": ["ROE", "ROA", "ROIC", "ROIC - WACC"],
              "DuPont": ["DuPont Nettomarge", "Kapitalumschlag", "EK-Multiplikator"],
              "Cashflow-Qualitaet": ["Cash Conversion (CFO/NI)", "FCF/NI", "Accruals-Ratio"],
              "Verschuldung & Solvenz": ["Nettoverschuldung/EBITDA", "Zinsdeckung", "Verschuldungsgrad", "Current Ratio"]}
    mult_idx = {"Kapitalumschlag", "EK-Multiplikator", "Nettoverschuldung/EBITDA", "Zinsdeckung", "Verschuldungsgrad", "Current Ratio"}
    rat, row = main["ratios"], 3
    for grp, items in groups.items():
        section(ws, row, chr(65 + len(yrs)), grp)
        for c, y in enumerate(yrs, 2): C(ws, f"{chr(64+c)}{row}", str(y), b=True, fill=SEC, al="right")
        row += 1
        for name in items:
            if name not in rat.index: continue
            C(ws, f"A{row}", name, bd=True)
            for c, v in enumerate(rat.loc[name].values, 2):
                C(ws, f"{chr(64+c)}{row}", _n(v), fmt=(MUL if name in mult_idx else PCT), al="right", bd=True)
            row += 1
        row += 1

# --------------------------------------------------------------- Bewertung (+ Sensitivitaet)
def _valuation(ws, main, models, blended, consensus, wacc, ccy):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 30
    for col in "BCDEFG": ws.column_dimensions[col].width = 13
    title(ws, "A1", "G1", f"BEWERTUNG (je Aktie in {ccy})")
    m, d = main["multiples"], main["dcf"]; price = main["latest"]["price"]
    nd = (_arr(main["financials"], "Nettoverschuldung")[-1]) * MM if _arr(main["financials"], "Nettoverschuldung") is not None else np.nan
    sh = main["latest"]["shares"]; base = d.get("base_fcf", np.nan)
    g0 = d.get("growth", np.nan); term = d.get("terminal", 0.025); w0 = d.get("wacc", wacc)

    section(ws, 3, "B", "Multiplikatoren"); row = 4
    for lab, v, fmt in [("KGV", m["pe"], MUL), ("EV/EBIT", m["ev_ebit"], MUL), ("EV/EBITDA", m["ev_ebitda"], MUL),
                        ("EV/FCF", m["ev_fcf"], MUL), ("KBV", m["pb"], MUL), ("FCF-Rendite", m["fcf_yield"], PCT),
                        ("Dividendenrendite", m["div_yield"], PCT)]:
        C(ws, f"A{row}", lab, bd=True); C(ws, f"B{row}", _n(v), fmt=fmt, al="right", bd=True); row += 1

    section(ws, row + 1, "G", "DCF (Free Cashflow to Firm)"); row += 2
    for lab, v, fmt in [("Basis-FCF (Mio.)", base / MM if base and not np.isnan(base) else np.nan, CUR),
                        ("Wachstum (J1-5)", g0, PCT), ("Terminales Wachstum", term, PCT), ("WACC", w0, PCT)]:
        C(ws, f"A{row}", lab, bd=True); C(ws, f"B{row}", _n(v), fmt=fmt, fill=YEL, al="right", bd=True); row += 1
    if base and not np.isnan(base) and not np.isnan(g0) and not np.isnan(w0):
        C(ws, f"A{row}", "Jahr", b=True, fill=LITE, bd=True)
        for k in range(1, 6): C(ws, f"{chr(65+k)}{row}", f"J{k}", b=True, fill=LITE, al="right", bd=True)
        row += 1
        proj = [base * (1 + g0) ** k for k in range(1, 6)]; pv = [proj[k - 1] / (1 + w0) ** k for k in range(1, 6)]
        C(ws, f"A{row}", "FCF (Mio.)", bd=True)
        for k in range(5): C(ws, f"{chr(66+k)}{row}", _n(proj[k] / MM), fmt=CUR, al="right", bd=True)
        row += 1
        C(ws, f"A{row}", "Barwert (Mio.)", bd=True)
        for k in range(5): C(ws, f"{chr(66+k)}{row}", _n(pv[k] / MM), fmt=CUR, al="right", bd=True)
        row += 1
    for lab, v, fmt in [("Enterprise Value (Mio.)", d.get("ev_dcf", np.nan) / MM if d.get("ev_dcf") else np.nan, CUR),
                        ("Eigenkapitalwert (Mio.)", d.get("equity_value", np.nan) / MM if d.get("equity_value") else np.nan, CUR),
                        ("Fair Value je Aktie", d.get("fair_value", np.nan), PS), ("Sicherheitsmarge", d.get("mos", np.nan), PCT),
                        ("Reverse-DCF impl. Wachstum", main["reverse_growth"], PCT)]:
        bold = lab.startswith("Fair") or lab.startswith("Sicher")
        C(ws, f"A{row}", lab, b=bold, bd=True)
        C(ws, f"B{row}", _n(v), b=bold, fmt=fmt, al="right", fill=(SEC if bold else None), bd=True); row += 1

    # Sensitivitaet WACC x Wachstum
    if base and not np.isnan(base) and not np.isnan(sh) and sh and not np.isnan(nd):
        section(ws, row + 1, "G", "DCF-Sensitivitaet: Fair Value je Aktie (WACC x Wachstum)"); row += 2
        gs = [g0 - 0.02, g0 - 0.01, g0, g0 + 0.01, g0 + 0.02]
        ws_ = [w0 - 0.01, w0 - 0.005, w0, w0 + 0.005, w0 + 0.01]
        C(ws, f"A{row}", "WACC \\ g", b=True, fill=LITE, al="center", bd=True)
        for j, gg in enumerate(gs, 2): C(ws, f"{chr(64+j)}{row}", _n(gg), b=True, fill=LITE, fmt=PCT, al="right", bd=True)
        row += 1; first = row
        for wv in ws_:
            C(ws, f"A{row}", _n(wv), b=True, fill=LITE, fmt=PCT, al="right", bd=True)
            for j, gg in enumerate(gs, 2):
                if wv > term:
                    proj = [base * (1 + gg) ** k for k in range(1, 6)]
                    pvs = sum(proj[k - 1] / (1 + wv) ** k for k in range(1, 6))
                    tv = proj[-1] * (1 + term) / (wv - term) / (1 + wv) ** 5
                    fv = (pvs + tv - nd) / sh
                else:
                    fv = np.nan
                C(ws, f"{chr(64+j)}{row}", _n(fv), fmt=PS, al="right", bd=True)
            row += 1
        rng = f"B{first}:F{row-1}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFC7CE", mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="C6EFCE"))
        C(ws, f"A{row}", "Farbskala: gruen = hoeherer Fair Value. Mittelzelle = Basisannahmen.", col="5A5A5A", sz=9)
        row += 1

    section(ws, row + 1, "C", "Methoden-Gegenueberstellung"); row += 2
    C(ws, f"A{row}", "Methode", b=True, fill=LITE, bd=True); C(ws, f"B{row}", "Fairer Wert", b=True, fill=LITE, al="right", bd=True)
    C(ws, f"C{row}", "Auf-/Abschlag", b=True, fill=LITE, al="right", bd=True); row += 1
    for lab, v in models.items():
        if v is None or np.isnan(v): continue
        up = (v / price - 1) if price and not np.isnan(price) else np.nan
        C(ws, f"A{row}", lab, bd=True); C(ws, f"B{row}", _n(v), fmt=PS, al="right", bd=True)
        c = C(ws, f"C{row}", _n(up), fmt=UP, al="right", bd=True)
        if up is not None and not np.isnan(up): c.fill = GRN if up >= 0 else REDF
        row += 1
    C(ws, f"A{row}", "Median", b=True, bd=True); C(ws, f"B{row}", _n(blended), b=True, fmt=PS, al="right", bd=True); row += 1

    if consensus and consensus.get("target_mean"):
        section(ws, row + 1, "B", "Analysten-Kursziele (externe Referenz)"); rr = row + 2
        for lab, v, fmt in [("Kursziel Mittel", consensus.get("target_mean"), PS), ("Kursziel Hoch", consensus.get("target_high"), PS),
                            ("Kursziel Tief", consensus.get("target_low"), PS), ("Anzahl Analysten", consensus.get("n_analysts"), "0")]:
            C(ws, f"A{rr}", lab, bd=True); C(ws, f"B{rr}", _n(v), fmt=fmt, al="right", bd=True); rr += 1

# --------------------------------------------------------------- Dividende
def _dividends(ws, main, ccy):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 26
    for col in "BCD": ws.column_dimensions[col].width = 14
    title(ws, "A1", "D1", "DIVIDENDE")
    di = main.get("div_info") or {}; dh = main.get("dividends") or []
    dy = di.get("yield_")
    if dy is not None and dy > 0.5: dy = dy / 100
    section(ws, 3, "D", "Kennzahlen"); row = 4
    for lab, v, fmt in [("Dividendenrendite", dy, PCT), ("Ausschuettungsquote", di.get("payout"), PCT),
                        ("Letzter Ex-Tag", di.get("ex_date"), None), ("DPS (letztes GJ)", main["latest"].get("dps"), PS)]:
        C(ws, f"A{row}", lab, bd=True)
        if fmt and isinstance(v, (int, float)): C(ws, f"B{row}", _n(v), fmt=fmt, al="right", bd=True)
        else: C(ws, f"B{row}", v or "n/v", al="right", bd=True)
        row += 1
    if dh:
        by = {}
        for x in dh: by[x["year"]] = by.get(x["year"], 0) + x["amount"]
        yrs = sorted(by)[-4:]; vals = [by[y] for y in yrs]
        growth = [None] + [(vals[i] / vals[i - 1] - 1) if vals[i - 1] else None for i in range(1, len(vals))]
        cagr = ((vals[-1] / vals[0]) ** (1 / (len(vals) - 1)) - 1) if len(vals) > 1 and vals[0] else np.nan
        section(ws, row + 1, "D", "Dividende je Aktie & Wachstum"); row += 2
        hr = row
        C(ws, f"A{row}", "Jahr", b=True, fill=LITE, bd=True); C(ws, f"B{row}", "DPS", b=True, fill=LITE, al="right", bd=True)
        C(ws, f"C{row}", "Wachstum", b=True, fill=LITE, al="right", bd=True); row += 1
        for i, y in enumerate(yrs):
            C(ws, f"A{row}", str(y), bd=True); C(ws, f"B{row}", _n(vals[i]), fmt=PS, al="right", bd=True)
            C(ws, f"C{row}", _n(growth[i]), fmt=UP, al="right", bd=True); row += 1
        C(ws, f"A{row}", "CAGR (Zeitraum)", b=True, bd=True); C(ws, f"C{row}", _n(cagr), b=True, fmt=PCT, al="right", bd=True)
        ch = BarChart(); ch.title = f"Dividende je Aktie ({ccy})"; ch.height = 6.5; ch.width = 11; ch.legend = None
        ch.add_data(Reference(ws, min_col=2, min_row=hr + 1, max_row=row - 1)); ch.set_categories(Reference(ws, min_col=1, min_row=hr + 1, max_row=row - 1))
        ws.add_chart(ch, "E3")
        section(ws, row + 2, "B", "Letzte Ausschuettungen (Ex-Tag)"); row += 3
        C(ws, f"A{row}", "Ex-Tag", b=True, fill=LITE, bd=True); C(ws, f"B{row}", "Betrag", b=True, fill=LITE, al="right", bd=True); row += 1
        for x in dh[::-1][:12]:
            C(ws, f"A{row}", x["date"], bd=True); C(ws, f"B{row}", _n(x["amount"]), fmt=PS, al="right", bd=True); row += 1
    else:
        C(ws, f"A{row+1}", "Keine Dividendenhistorie verfuegbar.", col="5A5A5A", sz=9)

# --------------------------------------------------------------- Qualitaet
def _quality(ws, main):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 34
    for col in "BCD": ws.column_dimensions[col].width = 13
    title(ws, "A1", "D1", "QUALITAETS-SCORECARD")
    labels = dict(geschaeftsmodell="Geschaeftsmodell*", management="Management*", wachstum="Wachstum",
                  profitabilitaet="Profitabilitaet (ROIC>WACC)", bilanz="Bilanz & Verschuldung",
                  cashflow="Cashflow-Qualitaet", margen="Margen", bewertung="Bewertung (Niveau)")
    C(ws, "A3", "Kriterium", b=True, fill=SEC, bd=True); C(ws, "B3", "Gewicht", b=True, fill=SEC, al="right", bd=True)
    C(ws, "C3", "Score (1-5)", b=True, fill=SEC, al="right", bd=True); C(ws, "D3", "Beitrag", b=True, fill=SEC, al="right", bd=True)
    row = 4
    for k in main["weights"]:
        C(ws, f"A{row}", labels[k], bd=True); C(ws, f"B{row}", main["weights"][k], fmt=PCT, al="right", bd=True)
        C(ws, f"C{row}", main["scores"][k], al="right", bd=True)
        C(ws, f"D{row}", round(main["weights"][k] * main["scores"][k], 2), fmt="0.00", al="right", bd=True); row += 1
    C(ws, f"A{row}", "Qualitaets-Score (gewichtet)", b=True, bd=True)
    C(ws, f"D{row}", round(main["conviction"], 2), b=True, fmt='0.00" / 5"', al="right", bd=True)
    C(ws, f"A{row+2}", "Der Qualitaets-Score bewertet ausschliesslich fundamentale Merkmale. Einordnung der "
                       "Unternehmensqualitaet, NICHT Kauf-/Verkaufsempfehlung. * qualitativ gesetzt.",
      col="5A5A5A", sz=9, wrap=True); ws.merge_cells(f"A{row+2}:D{row+4}")

# --------------------------------------------------------------- Peers
def _peers(ws, main, peer_rs, ccy):
    ws.sheet_view.showGridLines = False; ws.column_dimensions["A"].width = 24
    cols = [("Qual.-Score", "0.0"), ("Auf-/Abschlag", UP), ("ROIC", PCT), ("ROIC-WACC", PCT), ("EBIT-Marge", PCT),
            ("Umsatz-CAGR", PCT), ("ND/EBITDA", MUL), ("KGV", MUL), ("EV/EBIT", MUL), ("EV/EBITDA", MUL),
            ("FCF-Rendite", PCT), ("Div.-Rendite", PCT)]
    for i in range(len(cols)): ws.column_dimensions[chr(66 + i)].width = 12
    title(ws, "A1", f"{chr(65+len(cols))}1", "PEER-VERGLEICH (waehrungsneutrale Kennzahlen)")
    C(ws, "A2", "Titel", b=True, fill=SEC, bd=True)
    for j, (hn, _) in enumerate(cols, 2): C(ws, f"{chr(64+j)}2", hn, b=True, fill=SEC, al="right", bd=True)
    allr = [main] + peer_rs; row = 3; first = row
    for r in allr:
        hh, mm = r["headline"], r["multiples"]
        C(ws, f"A{row}", ("> " + r["ticker"]) if r is main else r["ticker"], b=(r is main), bd=True)
        vals = [_n(r["conviction"]), _n(hh["mos"]), _n(hh["roic"]), _n(hh["roic_wacc"]), _n(hh["ebit_margin"]),
                _n(hh["rev_cagr"]), _n(hh["nd_ebitda"]), _n(mm["pe"]), _n(mm["ev_ebit"]), _n(mm["ev_ebitda"]),
                _n(mm["fcf_yield"]), _n(mm["div_yield"])]
        for j, (v, (_, fmt)) in enumerate(zip(vals, cols), 2):
            C(ws, f"{chr(64+j)}{row}", v, fmt=fmt, al="right", bd=True)
        row += 1
    # Median (Peers)
    C(ws, f"A{row}", "Median (Peers)", b=True, fill=LITE, bd=True)
    def med(key, mult=False):
        a = [(_n(r["multiples"][key]) if mult else _n(r["headline"][key])) for r in peer_rs]
        a = [x for x in a if x is not None]; return float(np.median(a)) if a else None
    medvals = [float(np.median([r["conviction"] for r in peer_rs])), med("mos"), med("roic"), med("roic_wacc"),
               med("ebit_margin"), med("rev_cagr"), med("nd_ebitda"), med("pe", True), med("ev_ebit", True),
               med("ev_ebitda", True), med("fcf_yield", True), med("div_yield", True)]
    for j, (v, (_, fmt)) in enumerate(zip(medvals, cols), 2):
        C(ws, f"{chr(64+j)}{row}", v, b=True, fill=LITE, fmt=fmt, al="right", bd=True)

    # Chart KGV/EV-EBIT je Titel
    ch = BarChart(); ch.type = "col"; ch.title = "Bewertung je Titel (KGV / EV-EBIT)"; ch.height = 8; ch.width = 16
    pe_col, ev_col = 9, 10  # H, I (1-based: A=1..) -> KGV col index 9, EV/EBIT 10
    ch.add_data(Reference(ws, min_col=pe_col, min_row=2, max_row=first + len(allr) - 1), titles_from_data=True)
    ch.add_data(Reference(ws, min_col=ev_col, min_row=2, max_row=first + len(allr) - 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=first + len(allr) - 1))
    ws.add_chart(ch, f"A{row+3}")

def build_excel(main, peer_rs, models, blended, consensus, wacc):
    ccy = main.get("price_ccy") or main.get("currency") or ""
    wb = Workbook()
    _overview(wb.active, main, models, blended, consensus, ccy); wb.active.title = "Uebersicht"
    _statements(wb.create_sheet("Abschluesse"), main, ccy)
    _ratios(wb.create_sheet("Kennzahlen"), main)
    _valuation(wb.create_sheet("Bewertung"), main, models, blended, consensus, wacc, ccy)
    _dividends(wb.create_sheet("Dividende"), main, ccy)
    _quality(wb.create_sheet("Qualitaet"), main)
    if peer_rs: _peers(wb.create_sheet("Peers"), main, peer_rs, ccy)
    bio = BytesIO(); wb.save(bio); return bio.getvalue()

if __name__ == "__main__":
    import equity_engine as eng
    r = eng.compute(eng._synthetic())
    r["dividends"] = [{"date": f"{y}-08-15", "year": y, "amount": a} for y, a in [(2022, 1.0), (2023, 1.1), (2024, 1.25), (2025, 1.4)]]
    r["div_info"] = {"yield_": 0.72, "payout": 0.28, "ex_date": "2025-05-15"}
    peer = eng.compute(eng._synthetic()); peer["ticker"] = "PEER1"
    models = {"DCF (FCFF)": r["headline"]["fair_value"], "DDM (Gordon)": 22.9, "EV/EBIT (Peer)": 49.4, "Analysten-Ziel": 42.0}
    cons = dict(target_mean=42.0, target_high=55.0, target_low=30.0, n_analysts=14)
    open("selftest_report.xlsx", "wb").write(build_excel(r, [peer], models, 33.4, cons, 0.08)); print("Excel ok")
